import io, re, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
YELLOW="FFCB05"; INK="1A1A1A"; GREY="F7F7F5"
RAG_FILL={"green":"E8F5E9","amber":"FFF6E0","red":"FDECEC","blue":"EAF2FC"}
RAG_FONT={"green":"1B5E20","amber":"7A5A00","red":"B71C1C","blue":"0B3D91"}
THIN=Side(style="thin",color="D9D9D9"); BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
def rag_key(v):
    t=(v or "").strip().lower()
    if t.startswith("green") or "on track" in t: return "green"
    if t.startswith("amber") or "delay" in t: return "amber"
    if t.startswith("red") or "at risk" in t or "off track" in t: return "red"
    if t.startswith("blue") or "deliver" in t or "bau" in t: return "blue"
    return None
def _sn(t):
    m={"MI Initiatives":"MI Initiatives","Accelerate Initiatives":"Accelerate","MI & Accelerate Priorities (KPIs)":"Priorities (KPIs)","Monthly Actions Tracker":"Actions Tracker"}
    return re.sub(r"[\\/?*\[\]:]","",m.get(t,t))[:31]
def _w(h):
    t=(h or "").lower()
    if t=="section": return 22
    if t=="initiative": return 30
    if t in ("changed",) or "rag" in t or "slide" in t: return 12
    if "maturity" in t or "due" in t or "date" in t or "timeline" in t: return 16
    if any(k in t for k in ("impact","performance","ytd","kpi","bud","actual")): return 18
    return 42
def build_workbook(payload):
    wb=Workbook(); ws=wb.active; ws.title="Summary"; ws.sheet_view.showGridLines=False
    ws["A1"]="MTN EBU — Accelerate & Maturity Index"; ws["A1"].font=Font(name="Segoe UI",bold=True,size=16,color="0B0B0B")
    ws["A2"]="Monthly OpCo submission · Group EBU Strategy & Transformation"; ws["A2"].font=Font(name="Segoe UI",size=9,color="6B6B6B")
    meta=[("OpCo",payload.get("opco","")),("Reporting month",payload.get("reportingMonth","")),("Submitted by",payload.get("submittedBy","")),("Email",payload.get("email","")),("Submitted at",payload.get("submittedAt",datetime.datetime.utcnow().isoformat())),("Initiatives updated this month",str(payload.get("itemsUpdated","")))]
    r=4
    for k,v in meta:
        ws.cell(row=r,column=1,value=k).font=Font(name="Segoe UI",bold=True,color=INK); ws.cell(row=r,column=2,value=v); r+=1
    r+=1; ws.cell(row=r,column=1,value="Contents").font=Font(name="Segoe UI",bold=True); r+=1
    for c,h in enumerate(["Sheet","Source slide(s)","Initiatives","Updated"],start=1):
        cell=ws.cell(row=r,column=c,value=h); cell.font=Font(name="Segoe UI",bold=True,color="0B0B0B"); cell.fill=PatternFill("solid",fgColor=YELLOW); cell.border=BORDER
    cstart=r+1
    for w,col in zip([34,26,16,14],"ABCD"): ws.column_dimensions[col].width=w
    bt={}
    for s in payload.get("sections",[]):
        t=s.get("type","Section"); g=bt.setdefault(t,{"slides":[],"items":[]}); g["slides"].append(s.get("sourceSlide"))
        for it in s.get("items",[]):
            item=dict(it); item["__slide"]=s.get("sourceSlide"); g["items"].append(item)
    crows=[]
    for t,g in bt.items():
        if not g["items"]: continue
        fields=[]
        for it in g["items"]:
            for f in (it.get("fields") or {}):
                if f and f not in fields: fields.append(f)
        cm="Comment / risk"
        if cm in fields: fields.remove(cm)
        headers=["Section","Initiative"]+fields+[cm,"Changed","Slide"]
        sh=wb.create_sheet(_sn(t)); sh.sheet_view.showGridLines=False
        sh["A1"]="MTN EBU — "+t; sh["A1"].font=Font(name="Segoe UI",bold=True,size=16,color="0B0B0B")
        sh["A2"]="%s  ·  %s  ·  submitted by %s"%(payload.get("opco",""),payload.get("reportingMonth",""),payload.get("submittedBy","")); sh["A2"].font=Font(name="Segoe UI",size=9,color="6B6B6B")
        for c,h in enumerate(headers,start=1):
            cell=sh.cell(row=4,column=c,value=h); cell.font=Font(name="Segoe UI",bold=True,color="0B0B0B"); cell.fill=PatternFill("solid",fgColor=YELLOW); cell.border=BORDER; cell.alignment=Alignment(vertical="center",wrap_text=True); sh.column_dimensions[get_column_letter(c)].width=_w(h)
        updated=0; rr=5
        for i,it in enumerate(g["items"]):
            ch=bool(it.get("changed"))
            if ch: updated+=1
            bf=PatternFill("solid",fgColor=GREY) if (i%2==1) else None; fl=it.get("fields") or {}
            vals=[it.get("section",""),it.get("initiative","")]+[fl.get(f,"") for f in fields]+[fl.get(cm,""),"Yes" if ch else "No",it.get("__slide","")]
            for c,val in enumerate(vals,start=1):
                cell=sh.cell(row=rr,column=c,value=val); cell.border=BORDER; cell.alignment=Alignment(vertical="top",wrap_text=True); hdr=headers[c-1]
                if bf: cell.fill=bf
                if c==2: cell.font=Font(name="Segoe UI",bold=True,color=INK)
                if "rag" in hdr.lower():
                    k=rag_key(str(val))
                    if k: cell.fill=PatternFill("solid",fgColor=RAG_FILL[k]); cell.font=Font(name="Segoe UI",color=RAG_FONT[k])
                if hdr=="Changed" and ch: cell.fill=PatternFill("solid",fgColor=RAG_FILL["amber"])
            rr+=1
        sh.freeze_panes="A5"; sh.auto_filter.ref="A4:%s%d"%(get_column_letter(len(headers)),rr-1)
        crows.append((_sn(t),", ".join(str(x) for x in g["slides"] if x is not None),len(g["items"]),updated))
    for idx,(n,sl,ct,up) in enumerate(crows):
        for c,val in enumerate([n,sl,ct,up],start=1): ws.cell(row=cstart+idx,column=c,value=val).border=BORDER
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()
