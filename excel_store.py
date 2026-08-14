"""
Excel storage (OneDrive via Power Automate) for the DGW MI & Accelerate app.

On submit, the app builds one flat row per initiative and POSTs them to your Power
Automate flow, which appends them to the single master workbook (submissions.xlsx)
in YOUR OneDrive/SharePoint 'MI and Accelerate' folder. Submitters never touch the file.

Set the flow URL in st.secrets:  flow_url = "https://prod-..../invoke?..."
(or the FLOW_URL environment variable). With no flow URL, it falls back to a local
submissions.xlsx so you can test on your PC.
"""

import io
import os
import datetime

import common

LOCAL_PATH = os.environ.get("MASTER_XLSX",
                            os.path.join(os.path.dirname(__file__), "submissions.xlsx"))

# must match the columns/table in the pre-made submissions.xlsx
COLUMNS = ["OpCo", "Reporting month", "Submitted by", "Email", "Type", "Section",
           "Initiative", "RAG", "Accelerate %", "Actual", "Estimated",
           "Maturity %", "Current", "Target", "Comment", "Submitted UTC",
           "Submitted date", "Submitted time"]


def get_flow_url():
    try:
        import streamlit as st
        u = st.secrets.get("flow_url", "")
        if u:
            return str(u).strip()
    except Exception:
        pass
    return os.environ.get("FLOW_URL", "").strip()


def _split_ts(iso):
    s = (iso or "").replace("Z", "")
    d = t = ""
    if "T" in s:
        d, t = s.split("T", 1)
        t = t.split(".")[0]
    return iso or "", d, t


def _payload_to_rows(payload):
    iso, dpart, tpart = _split_ts(payload.get("submittedAt", ""))
    rows = []
    for rec in common.submissions_to_records([payload]):
        f = rec["fields"]
        ap = act = est = mp = cur = tgt = ""
        if rec["kind"] == "Accelerate":
            p, a, e, _ek, _ak = common.accel_progress(f)
            ap = round(p, 1) if p is not None else ""
            act = a if a is not None else ""
            est = e if e is not None else ""
        if rec["kind"] == "MI":
            p, cu, tg = common.maturity_progress(f)
            mp = round(p, 0) if p is not None else ""
            cur, tgt = cu or "", tg or ""
        rag = common.rag_bucket(f.get("RAG Status", ""))
        rows.append({
            "OpCo": rec["opco"], "Reporting month": payload.get("reportingMonth", ""),
            "Submitted by": payload.get("submittedBy", ""), "Email": payload.get("email", ""),
            "Type": rec["type"], "Section": rec.get("section", ""), "Initiative": rec["initiative"],
            "RAG": rag, "Accelerate %": ap, "Actual": act, "Estimated": est,
            "Maturity %": mp, "Current": cur, "Target": tgt,
            "Comment": f.get("Comment / risk", ""),
            "Submitted UTC": iso, "Submitted date": dpart, "Submitted time": tpart,
        })
    return rows


def _post_to_flow(url, payload):
    import requests
    rows = _payload_to_rows(payload)
    body = {"opco": payload.get("opco", ""), "reportingMonth": payload.get("reportingMonth", ""),
            "submittedBy": payload.get("submittedBy", ""), "rows": rows}
    r = requests.post(url, json=body, timeout=45)
    r.raise_for_status()
    return len(rows)


def _local_save(payload):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    if os.path.exists(LOCAL_PATH):
        wb = load_workbook(LOCAL_PATH)
        ws = wb["Submissions"] if "Submissions" in wb.sheetnames else wb.active
    else:
        wb = Workbook(); ws = wb.active; ws.title = "Submissions"; ws.append(COLUMNS)
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True, color="0B0B0B")
            ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor="FFCB05")
    for row in _payload_to_rows(payload):
        ws.append([row.get(c, "") for c in COLUMNS])
    wb.save(LOCAL_PATH)


def _local_all_rows():
    from openpyxl import load_workbook
    if not os.path.exists(LOCAL_PATH):
        return []
    wb = load_workbook(LOCAL_PATH, data_only=True)
    ws = wb["Submissions"] if "Submissions" in wb.sheetnames else wb.active
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0] and str(r[0]).strip() and str(r[0]) != "(sample)":
            out.append(dict(zip(COLUMNS, r)))
    return out


# --------------------------------------------------------------------------- #
#  Public API
# --------------------------------------------------------------------------- #
def save_submission(payload, year, month):
    if not payload.get("submittedAt"):
        payload["submittedAt"] = datetime.datetime.utcnow().isoformat() + "Z"
    url = get_flow_url()
    if url:
        return _post_to_flow(url, payload)     # -> appends to OneDrive master
    _local_save(payload)                       # -> local submissions.xlsx
    return len(_payload_to_rows(payload))


def storage_label():
    return "OneDrive (Power Automate)" if get_flow_url() else "local file · submissions.xlsx"


def load_submissions():
    """With the flow backend the master lives in OneDrive and is not read back here,
    so the in-app dashboard preview stays empty online. Locally, it reflects the file."""
    return []


def master_bytes():
    if os.path.exists(LOCAL_PATH):
        with open(LOCAL_PATH, "rb") as fh:
            return fh.read()
    from openpyxl import Workbook
    wb = Workbook(); wb.active.append(COLUMNS)
    b = io.BytesIO(); wb.save(b); return b.getvalue()
